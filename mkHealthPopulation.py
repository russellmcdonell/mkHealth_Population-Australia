#!/usr/bin/env python

# pylint: disable=invalid-name, line-too-long, pointless-string-statement, unused-variable, broad-exception-caught, too-many-lines

'''
A script to create an Excel spreadsheet of randomly created health networks, hospitals, clinics, specialists, GPs and patients

SYNOPSIS
$ python mkHealthPopulation.py [-D dataDir|--dataDir=dataDir] [-A addressFile|--addressFile=addressFile]
                               [-O outputDir|--outputDir=outputDir] [-o outputfile|--outputfile=outputfile]
                               [-r|--makeRandom] [-P|--Patients] [-i|--IHI] [-x|--extendNames] [-a|--addUR]
                               [-v loggingLevel|--loggingLevel=loggingLevel]
                               [-L logDir|--logDir=logDir] [-l logfile|--logfile=logfile]

OPTIONS
-D dataDir|--dataDir=dataDir
The directory containing the source names and address data (default='data')

-A addressFile|--addressFile=addressFile
The file of GNAF_CORE addresses (or subset) (default='GNAF_CORE.psv')

-O outputDir|--outputDir=outputDir
The directory in which the output file will be created (default='output')
[mkHealthPopulation.cfg will be read from this directory]

-o outputfile|--outfile=outputfile
The output file to be created (default='clinicDoctors.csv')

-r|--makeRandom
Make up Australian addresses by random selection of street, suburb, state, postcode

-P|--Patients
Output Patients for each GP doctor

-i|--IHI
Add Australian IHI number

-x|--extendName
Extend names with sequential letters

-a|--addUR
Add a template for the UR number  in the PID and LIS2 segments

-v loggingLevel|--verbose=loggingLevel
Set the level of logging that you want.

-L logDir|--logDir=logDir
The name of the folder for the logging file

-l logfile|--logfile=logfile
The name of a logging file where you want all messages captured.
'''

import sys
import os
import csv
import argparse
import logging
import configparser
import random
import json
from openpyxl import Workbook
from fhir.resources.organization import Organization
from fhir.resources.healthcareservice import HealthcareService
from fhir.resources.location import Location
from fhir.resources.practitioner import Practitioner
from fhir.resources.practitionerrole import PractitionerRole
from fhir.resources.careteam import CareTeam
from fhir.resources.patient import Patient
from randPatients import patients, patientKeys, mkRandPatients, mkRandAddress, mkLuhn, SA3postcodes, SA2inSA4, SA1s


careTeams = {}        # The list of Practitioners in each Organization
patientDetails = {}    # The patient details plus the list of GPs

# This next section is plagurised from /usr/include/sysexits.h
EX_OK = 0                # successful termination
EX_WARN = 1                # non-fatal termination with warnings

EX_USAGE = 64            # command line usage error
EX_DATAERR = 65            # data format error
EX_NOINPUT = 66            # cannot open input
EX_NOUSER = 67            # addressee unknown
EX_NOHOST = 68            # host name unknown
EX_UNAVAILABLE = 69        # service unavailable
EX_SOFTWARE = 70        # internal software error
EX_OSERR = 71            # system error (e.g., can't fork)
EX_OSFILE = 72            # critical OS file missing
EX_CANTCREAT = 73        # can't create (user) output file
EX_IOERR = 74            # input/output error
EX_TEMPFAIL = 75        # temp failure; user is invited to retry
EX_PROTOCOL = 76        # remote error in protocol
EX_NOPERM = 77            # permission denied
EX_CONFIG = 78            # configuration error


def mkProviderNo(thisProviderNo):
    '''
    Add a provider location and checksum to a provider number
    '''
    weights = [3, 5, 8, 4, 2, 1]
    providerLocation = '0123456789ABCDEFGHJKLMNPQRTUVWXY'
    csumChar = 'YXWTLKJHFBA'

    csum = 0
    for i, weight in enumerate(weights):
        csum += int(thisProviderNo[i]) * weight
        providerLoc = random.choice(range(len(providerLocation)))
        csum += providerLoc * 6
    csum %= 11
    return f'{thisProviderNo}{providerLocation[providerLoc:providerLoc+1]}{csumChar[csum:csum+1]}'


def mkPrescriberNo(prescriberNo):
    '''
    Add a check digit to a PBS prescriber number
    '''
    cdigit = '0123456789'
    csum = 0
    if prescriberNo[0] == '0':
        csum = (int(prescriberNo[1]) * 5 + int(prescriberNo[2]) * 8 + int(prescriberNo[3]) * 4 + int(prescriberNo[4]) * 2 + int(prescriberNo[5])) % 11
    else:
        csum = int(prescriberNo[0]) + int(prescriberNo[1]) * 3 + int(prescriberNo[2]) * 7 + int(prescriberNo[3]) * 9 + int(prescriberNo[4]) + int(prescriberNo[5]) * 3
    csum %= 10
    return prescriberNo + cdigit[csum]


def createOrganization(orgRecord, HPIO, partOfHPIO, orgName, addr, organizationType):
    '''
    Create the FHIR_Organization resource
    '''

    logging.info('Creating organization %s', orgName)

    rowData = patients[patientKeys[orgRecord]]
    organization_dict = {'resourceType':'Organization',
        'id': HPIO}
    if addr is None:
        organization_dict.update({
            'text': {
                'status': 'generated',
                'div': f"<div xmlns='http://www.w3.org/1999/xhtml'><table><tbody><tr><td>Name</td><td>{orgName}</td></tr>" + \
                    f"<tr><td>Address</td><td>{rowData['streetNo']} {rowData['streetName']} {rowData['shortStreetType']}</td></tr>" + \
                    f"<tr><td></td><td>{rowData['suburb']}</td></tr><tr><td></td><td>{rowData['state']}</td></tr><tr><td></td><td>{rowData['postcode']}</td></tr></tbody></table></div>"
            }
        })
    else:
        organization_dict.update({
            'text': {
                'status': 'generated',
                'div': f"<div xmlns='http://www.w3.org/1999/xhtml'><table><tbody><tr><td>Name</td><td>{orgName}</td></tr>" + \
                    f"<tr><td>Address</td><td>{addr['streetNo']} {addr['streetName']} {addr['shortStreetType']}</td></tr>" + \
                    f"<tr><td></td><td>{addr['suburb']}</td></tr><tr><td></td><td>{addr['state']}</td></tr><tr><td></td><td>{addr['postcode']}</td></tr></tbody></table></div>"
            }
        })
    organization_dict.update({
        'identifier': [
            {
                'use':'official',
                'type': {
                    'coding': [
                        {
                            'system': 'http://terminology.hl7.org.au/CodeSystem/v2-0203',
                            'code': 'NOI',
                            'display': 'National Organisation Identifier'
                        }
                    ],
                    'text': 'HPI-O'
                },
                'system': 'http://ns.electronichealth.net.au/id/hi/hpio/1.0',
                'value': HPIO
            }
        ],
        'active': True
    })
    if organizationType == 'public hospital':
        organization_dict.update({
            'type': [
                        {
                        'coding': [
                            {
                                'system': 'http://snomed.info/sct',
                                'code': '257622000',
                                'display': HealthcareRoles['257622000']
                            }
                        ]
                    }
                ],
            'name': orgName
        })
    elif organizationType == 'private hospital':
        organization_dict.update({
            'type': [
                        {
                        'coding': [
                            {
                                'system': 'http://snomed.info/sct',
                                'code': '309895006',
                                'display': HealthcareRoles['309895006']
                            }
                        ]
                    }
                ],
            'name': orgName
        })
    elif organizationType == 'department':
        organization_dict.update({
            'type': [
                        {
                        'coding': [
                            {
                                'system': 'http://snomed.info/sct',
                                'code': departmentRoles[orgName],
                                'display': HealthcareRoles[departmentRoles[orgName]]
                            }
                        ]
                    }
                ],
            'name': orgName
        })
    elif organizationType == 'clinic':
        organization_dict.update({
            'type': [
                        {
                        'coding': [
                            {
                                'system': 'http://snomed.info/sct',
                                'code': '288565001',
                                'display': HealthcareRoles['288565001']
                            }
                        ]
                    }
                ],
            'name': orgName
        })
    elif organizationType == 'specialist':
        organization_dict.update({
            'type': [
                        {
                        'coding': [
                            {
                                'system': 'http://snomed.info/sct',
                                'code': '83891005',
                                'display': HealthcareRoles['83891005']
                            }
                        ]
                    }
                ],
            'name': orgName
        })
    organization_dict.update({
        'telecom': [
            {
                'system': 'phone',
                'value': rowData['businessPhone'],
                'use': 'work'
            }
        ]
    })
    if addr is None:
        organization_dict.update({
            'address': [
                {
                    'use': 'work',
                    'type': 'postal',
                    'text': rowData['streetNo'] + ' ' + rowData['streetName'] + ' ' + rowData['shortStreetType'] + ', ' + rowData['suburb'] + ', ' + rowData['state'] + '     ' + rowData['postcode'],
                    'line': [rowData['streetNo'] + ' ' + rowData['streetName'] + ' ' + rowData['shortStreetType']],
                    'city': rowData['suburb'],
                    'state': rowData['state'],
                    'postalCode': rowData['postcode'],
                    'country': rowData['country']
                }
            ]
        })
    else:
        organization_dict.update({
            'address': [
                {
                    'use': 'work',
                    'type': 'postal',
                    'text': addr['streetNo'] + ' ' + addr['streetName'] + ' ' + addr['shortStreetType'] + ', ' + addr['suburb'] + ', ' + addr['state'] + '     ' + addr['postcode'],
                    'line': [addr['streetNo'] + ' ' + addr['streetName'] + ' ' + addr['shortStreetType']],
                    'city': addr['suburb'],
                    'state': addr['state'],
                    'postalCode': addr['postcode'],
                    'country': addr['country']
                }
            ]
        })
    if partOfHPIO is not None:
        organization_dict.update({
            'partOf': { 'reference': 'Organization/' + partOfHPIO }
        })

    organization = Organization.parse_obj(organization_dict)
    try:
        organization_json = json.dumps(organization_dict)
    except Exception as e:
        logging.critical('Bad JSON:%s', e.args)
        print(organization_dict)
        sys.stdout.flush()
        logging.shutdown()
        sys.exit(EX_DATAERR)
    # print(json.dumps(organization_dict, indent=4))

    # Return the resources
    if len(organization_json) > 32767:
        logging.critical('Maximum string length for an Excel cell exceeded')
        print(organization_dict)
        sys.stdout.flush()
        logging.shutdown()
        sys.exit(EX_DATAERR)

    return organization_json


def createLocation(locRecord, HPIO, locName, addr, managingHPIO):
    '''
    Create a FHIR_Location resource
    '''
    rowData = patients[patientKeys[locRecord]]
    location_dict = {'resourceType':'Location',
        'id': HPIO,
        'text': {
            'status': 'generated',
            'div': f"<div xmlns='http://www.w3.org/1999/xhtml'><table><tbody><tr><td>Name</td><td>{locName}</td></tr>" + \
                f"<tr><td>Address</td><td>{addr['streetNo']} {addr['streetName']} {addr['shortStreetType']}</td></tr>" + \
                f"<tr><td></td><td>{addr['suburb']}</td></tr><tr><td></td><td>{addr['state']}</td></tr><tr><td></td><td>{addr['postcode']}</td></tr></tbody></table></div>"
        },
        'status': 'active',
        'name': locName,
        'address': {
                'use': 'work',
                'type': 'postal',
                'text': addr['streetNo'] + ' ' + addr['streetName'] + ' ' + addr['shortStreetType'] + ', ' + addr['suburb'] + ', ' + addr['state'] + '     ' + addr['postcode'],
                'line': [addr['streetNo'] + ' ' + addr['streetName'] + ' ' + addr['shortStreetType']],
                'city': addr['suburb'],
                'state': addr['state'],
                'postalCode': addr['postcode'],
                'country': addr['country']
        },
        'position': {
                'longitude': float(addr['longitude']),
                'latitude': float(addr['latitude'])
        }
    }
    if managingHPIO is not None:
        location_dict.update({
            'managingOrganization': { 'reference': 'Organization/' + managingHPIO }
        })
    location = Location.parse_obj(location_dict)
    try:
        location_json = json.dumps(location_dict)
    except Exception as e:
        logging.critical('Bad JSON:%s', e.args)
        sys.stdout.flush()
        logging.shutdown()
        sys.exit(EX_DATAERR)
    # print(json.dumps(location_dict, indent=4))

    # Return the resources
    if len(location_json) > 32767:
        logging.critical('Maximum string length for an Excel cell exceeded')
        print(location_dict)
        sys.stdout.flush()
        logging.shutdown()
        sys.exit(EX_DATAERR)

    return location_json


def createHealthcareService(serviceRecord, HPIO, serviceName, serviceType, serviceSpecialty):
    '''
    Create the FHIR_HealthcareService resource
    '''
    rowData = patients[patientKeys[serviceRecord]]
    healthcareService_dict = {'resourceType':'HealthcareService',
        'id': HPIO,
        'text': {
            'status': 'generated',
            'div': f"<div xmlns='http://www.w3.org/1999/xhtml'><table><tbody><tr><td>Name</td><td>{serviceName}</td></tr>" + \
                f"<tr><td>Address</td><td>{rowData['streetNo']} {rowData['streetName']} {rowData['shortStreetType']}</td></tr>" + \
                f"<tr><td></td><td>{rowData['suburb']}</td></tr><tr><td></td><td>{rowData['state']}</td></tr><tr><td></td><td>{rowData['postcode']}</td></tr></tbody></table></div>"
        },
        'identifier': [
            {
                'use':'official',
                'type': {
                    'coding': [
                        {
                            'system': 'http://terminology.hl7.org.au/CodeSystem/v2-0203',
                            'code': 'NOI',
                            'display': 'National Organisation Identifier'
                        }
                    ],
                    'text': 'HPI-O'
                },
                'system': 'http://ns.electronichealth.net.au/id/hi/hpio/1.0',
                'value': HPIO
            }
        ],
        'active': True,
        'providedBy': {
                'reference': 'Organization/' + HPIO
        },
    }
    if serviceType == 'department':
        deptName = serviceName.split('@')[0]
        healthcareService_dict.update({
            'type': [
                {
                    'coding': [
                        {
                            'system': 'http://snomed.info/sct',
                            'code': departmentRoles[deptName],
                            'display': HealthcareRoles[departmentRoles[deptName]]
                        }
                    ]
                }
            ]
        })
    elif serviceType == 'clinic':
        healthcareService_dict.update({
            'type': [
                {
                    'coding': [
                        {
                            'system': 'http://snomed.info/sct',
                            'code': '288565001',
                            'display': HealthcareRoles['288565001']
                        }
                    ]
                }
            ]
        })
    elif serviceType == 'specialist':
        healthcareService_dict.update({
            'type': [
                {
                    'coding': [
                        {
                            'system': 'http://snomed.info/sct',
                            'code': '288565001',
                            'display': HealthcareRoles['288565001']
                        }
                    ]
                }
            ]
        })
    healthcareService_dict.update({
        'specialty': [
            {
                'coding': [
                    {
                        'system': 'http://snomed.info/sct',
                        'code': serviceSpecialty,
                        'display': Specialties[serviceSpecialty]
                    }
                ]
            }
        ],
        'location': [
            {
                'reference': 'Location/' + HPIO
            }
        ],
        'name': serviceName,
        'telecom': [
            {
                'system': 'phone',
                'value': rowData['businessPhone'],
                'use': 'work'
            }
        ]
    })

    healthcareService = HealthcareService.parse_obj(healthcareService_dict)
    try:
        healthcareService_json = json.dumps(healthcareService_dict)
    except Exception as e:
        logging.critical('Bad JSON:%s', e.args)
        sys.stdout.flush()
        logging.shutdown()
        sys.exit(EX_DATAERR)
    # print(json.dumps(healthcareService_dict, indent=4))

    # Return the resources
    if len(healthcareService_json) > 32767:
        logging.critical('Maximum string length for an Excel cell exceeded')
        print(healthcareService_dict)
        sys.stdout.flush()
        logging.shutdown()
        sys.exit(EX_DATAERR)

    return healthcareService_json


sexName = {'M': 'male', 'F': 'female', 'U': 'unknown'}
def createPractitioner(pracRecord, HPIO, HPII, pracProviderNo, pracAhpraNo):
    '''
    Create FHIR_Practitioner resource
    '''

    logging.info('Creating Practitioner :%s', pracProviderNo)

    rowData = patients[patientKeys[pracRecord]]
    if HPIO not in careTeams:
        careTeams[HPIO] = set()
    careTeams[HPIO].add(HPII)
    pracTitle = rowData['title']
    family = rowData['familyName']
    given = rowData['givenName']
    sex = sexName[rowData['sex']]
    practitioner_dict = {'resourceType':'Practitioner',
        'id': HPII,
        'text': {
            'status': 'generated',
            'div': f"<div xmlns='http://www.w3.org/1999/xhtml'><table><tbody><tr><td>Name</td><td>{pracTitle} {given} {family}</td></tr></tbody></table></div>"
        },
        'identifier': [
            {
                'type': {
                    'coding': [
                        {'system': 'http://terminology.hl7.org/CodeSystem/v2-0203',
                        'code': 'NPI',
                        'display': 'National Provider Identifier'}
                    ],
                    'text': 'HPI-I'
                },
                'system': 'http://ns.electronichealth.net.au/id/hi/hpii/1.0',
                'value': HPII
            },
            {
                'type': {
                    'coding': [
                        {'system': 'http://terminology.hl7.org/CodeSystem/v2-0203',
                        'code': 'PRES',
                        'display': 'Prescriber Number'}
                    ],
                    'text': 'Prescriber Number'
                },
                'system': 'http://ns.electronichealth.net.au/id/medicare-provider-number',
                'value': mkProviderNo(pracProviderNo)
            },
            {
                'type': {
                    'coding': [
                        {'system': 'http://terminology.hl7.org/CodeSystem/v2-0203',
                        'code': 'PRES',
                        'display': 'Prescriber Number'}
                    ],
                    'text': 'Prescriber Number'
                },
                'system': 'http://ns.electronichealth.net.au/id/medicare-prescriber-number',
                'value': mkPrescriberNo(pracProviderNo)
            },
            {
                'type': {
                    'coding': [
                        {'system': 'http://terminology.hl7.org/CodeSystem/v2-0203',
                        'code': 'PRES',
                        'display': 'Prescriber Number'}
                    ],
                    'text': 'Prescriber Number'
                },
                'system': 'http://hl7.or.au/id/ahpra-registration-number',
                'value': 'MED' + str(pracAhpraNo)
            }
        ],
        'active': True,
        'name': [
            {'use': 'official',
            'text': given + ' ' + family,
            'family': family,
            'given': [given],
            'prefix': [pracTitle]}
        ],
        'telecom': [
            {'system': 'phone',
            'value': rowData['businessPhone'],
            'use': 'work'},
            {'system': 'phone',
            'value': rowData['mobile'],
            'use': 'mobile'},
            {'system': 'email',
            'value': rowData['email'],
            'use': 'work'}
        ],
        'gender': sex,
        'birthDate': rowData['birthdate']
    }
    practitioner = Practitioner.parse_obj(practitioner_dict)
    try:
        practitioner_json = json.dumps(practitioner_dict)
    except Exception as e:
        logging.critical('Bad JSON:%s', e.args)
        sys.stdout.flush()
        logging.shutdown()
        sys.exit(EX_DATAERR)
    # print(json.dumps(practitioner_dict, indent=4))

    # Return the resources
    if len(practitioner_json) > 32767:
        logging.critical('Maximum string length for an Excel cell exceeded')
        print(practitioner_dict)
        sys.stdout.flush()
        logging.shutdown()
        sys.exit(EX_DATAERR)

    return practitioner_json


def createPractitionerRole(roleRecord, HPII, HPIO, roleRole, rolespecialty):
    '''
    Create FHIR_PractitionerRole resource
    '''
    rowData = patients[patientKeys[roleRecord]]
    roleTitle = rowData['title']
    family = rowData['familyName']
    given = rowData['givenName']
    practitionerRole_dict = {'resourceType':'PractitionerRole',
        'id': HPII + '-' + HPIO,
        'text': {
            'status': 'generated',
            'div': f"<div xmlns='http://www.w3.org/1999/xhtml'><table><tbody><tr><td>Name</td><td>{roleTitle} {given} {family}</td></tr></tbody></table></div>"
        },
        'active': True,
        'practitioner': {
            'reference': 'Practitioner/' + HPII
        },
        'organization': {
            'reference': 'Organization/' + HPIO
        },
        'code': [
            {
                'coding': [
                    {
                        'system': 'http://snomed.info/sct',
                        'code': roleRole,
                        'display': Roles[roleRole]
                    }
                ]
            }
        ],
        'specialty': [
            {
                'coding': [
                    {
                        'system': 'http://snomed.info/sct',
                        'code': rolespecialty,
                        'display': Specialties[rolespecialty]
                    }
                ]
            }
        ],
        'telecom': [
            {'system': 'phone',
            'value': rowData['businessPhone'],
            'use': 'work'},
            {'system': 'phone',
            'value': rowData['mobile'],
            'use': 'mobile'},
            {'system': 'email',
            'value': rowData['email'],
            'use': 'work'}
        ]
    }

    practitionerRole = PractitionerRole.parse_obj(practitionerRole_dict)
    try:
        practitionerRole_json = json.dumps(practitionerRole_dict)
    except Exception as e:
        logging.critical('Bad JSON:%s', e.args)
        sys.stdout.flush()
        logging.shutdown()
        sys.exit(EX_DATAERR)
    # print(json.dumps(practitionerRole_dict, indent=4))

    # Return the resources
    if len(practitionerRole_json) > 32767:
        logging.critical('Maximum string length for an Excel cell exceeded')
        print(practitionerRole_dict)
        sys.stdout.flush()
        logging.shutdown()
        sys.exit(EX_DATAERR)

    return practitionerRole_json


def createCareTeam(teamRecord, teamIHI, HPIO):
    '''
    Create the FHIR_CareTeam resource
    '''
    if len(careTeams[HPIO]) == 0:
        return
    if teamRecord is None:        # Template
        teamName = 'ReplaceWithGivenName ReplaceWithFamilyName'
    else:
        rowData = patients[patientKeys[teamRecord]]
        family = rowData['familyName']
        given = rowData['givenName']
        teamName = given + ' ' + family
    careteam_dict = {'resourceType':'CareTeam',
        'id': teamIHI + '-' + HPIO,
        'status': 'active',
        'name': teamName,
        'subject': {
            'reference': 'Patient/' + teamIHI
        },
        'participant': [
        ]
    }
    for teamMember in careTeams[HPIO]:
        careteam_dict['participant'].append({
            'member': {
                'reference': 'Practitioner/' + teamMember
            }
        })
    careteam_dict.update({
        'managingOrganization': [
            {'reference': 'Organization/' + HPIO }
        ]
    })

    careteam = CareTeam.parse_obj(careteam_dict)
    try:
        careteam_json = json.dumps(careteam_dict)
    except Exception as e:
        logging.critical('Bad JSON:%s', e.args)
        sys.stdout.flush()
        logging.shutdown()
        sys.exit(EX_DATAERR)
    # print(json.dumps(careteam_dict, indent=4))

    # Return the resources
    if len(careteam_json) > 32767:
        logging.critical('Maximum string length for an Excel cell exceeded')
        print(careteam_dict)
        sys.stdout.flush()
        logging.shutdown()
        sys.exit(EX_DATAERR)

    return careteam_json


marriedName = {'D': 'Divorced', 'M': 'Married', 'S': 'Never Married', 'W': 'Wodowed'}
raceName = {'1': 'Aboriginal but not Torres Straight Islander origin', '2': 'Torres Straight Islander but not Aboriginal origin',
            '3': 'Both Aboriginal and Torres Straight Islander origin', '4': 'Neither Aboriginal nor Torres Straight Islander origin',
            '9': 'Not stated/inadequately described'}
dvaName = {'GOL': 'Gold Card', 'WHT': 'White Card', 'ORN': 'Repatriation Pharaceutical Benefits Card'}
dvaCode = {'GOL': 'DVG', 'WHT': 'DVW', 'ORN': 'DVO'}
dvaColor = {'GOL': 'Gold', 'WHT': 'White', 'ORN': 'Orange'}
def createPatient(patientRecord, patientIHI, patientGPs):
    '''
    Create the FHIR_Patient resource
    '''

    logging.info('Creating Patient :%s', patientIHI)

    rowData = patients[patientKeys[patientRecord]]
    patientTitle = rowData['title']
    family = rowData['familyName']
    given = rowData['givenName']
    race = rowData['race']
    sex = sexName[rowData['sex']]
    birthdate = rowData['birthdate']
    born = birthdate.split('-')
    line = rowData['streetNo'] + ' ' + rowData['streetName'] + ' ' + rowData['shortStreetType']
    city = rowData['suburb']
    state = rowData['state']
    postcode = rowData['postcode']
    medicareNo = rowData['medicareNo']
    dvaNo = rowData['dvaNo']
    dvaType = rowData['dvaType']
    PEN = rowData['PEN']
    SEN = rowData['SEN']
    HC = rowData['HC']
    married = rowData['married']
    patient_dict = {
        'resourceType':'Patient',
        'id': patientIHI,
        'text': {
            'status': 'generated',
            'div': f"<div xmlns='http://www.w3.org/1999/xhtml'><table><tbody><tr><td>Name</td><td>{patientTitle} {given} {family}</td></tr>" + \
                    f"<tr><td>Born</td><td>{born[2]}/{born[1]}/{born[0]}</td></tr>" + \
                    f"<tr><td>Address</td><td>{line}</td></tr>" + \
                    f"<tr><td></td><td>{city}</td></tr><tr><td></td><td>{state}</td></tr><tr><td></td><td>{postcode}</td></tr></tbody></table></div>"
        } ,
        'extension': [
            {
                'url': 'http://hl7.org.au/fhir/StructureDefinition/indigenous-status',
                'valueCoding': {
                    'system': 'https://healthterminologies.gov.au/fhir/CodeSystem/australian-indigenous-status-1',
                    'code': race,
                    'display': raceName[race]
                }
            }
        ],
        'identifier': [
            {
                'extension': [
                    {
                        'url': 'http://hl7.org.au/fhir/StructureDefinition/ihi-status',
                        'valueCoding': {
                            'system': 'https://healthterminologies.gov.au/fhir/CodeSystem/ihi-status-1',
                            'code': 'Active',
                            'display': 'Active'
                        }
                    },
                    {
                        'url': 'http://hl7.org.au/fhir/StructureDefinition/ihi-record-status',
                        'valueCoding': {
                            'system': 'https://healthterminologies.gov.au/fhir/CodeSystem/ihi-record-status-1',
                            'code': 'Verified',
                            'display': 'Verified'
                        }
                    }
                ],
                'type': {
                    'coding': [
                        {
                            'system': 'http://hl7.org.au/fhir/v2/0203',
                            'code': 'NI',
                            'display': 'National Unique Individual Identifier'
                        }
                    ],
                    'text': 'IHI'
                },
                'system': 'http://ns.electronichealth.net.au/id/hi/ihi/1.0',
                'value': patientIHI
            }
        ],
        'active': True,
        'name': [
            {
                'use': 'official',
                'text': patientTitle + ' ' + given + ' ' + family,
                'family': family,
                'given': [given],
                'prefix': [patientTitle]
            }
        ],
        'telecom': [
            {
                'system': 'phone',
                'value': rowData['homePhone'],
                'use': 'home'
            }, {
                'system': 'phone',
                'value': rowData['businessPhone'],
                'use': 'work'
            }, {
                'system': 'phone',
                'value': rowData['mobile'],
                'use': 'mobile'
            }, {
                'system': 'email',
                'value': rowData['email'],
                'use': 'home'
            }
        ],
        'gender': sex,
        'birthDate': birthdate,
        'address': [
            {
                'use': 'home',
                'type': 'postal',
                'text': line + ', ' + city + ', ' + state + '     ' + postcode,
                'line': [line],
                'city': city,
                'state': state,
                'postalCode': postcode
            }
        ],
        'maritalStatus': {
            'coding': [
                {
                    'system': 'http://terminology.hl7.org/CodeSystem/v3-MaritalStatus',
                    'code': married,
                    'display': marriedName[married]
                }
            ],
            'text': marriedName[married]
        },
        'generalPractitioner': [
        ]
    }
    if medicareNo is not None:
        patient_dict['identifier'].append({
            'type': {
                'coding': [
                    {'system': 'http://terminology.hl7.org/CodeSystem/v2-0203',
                    'code': 'MC',
                    'display': "Patient's Medicare Number"}
                ],
                'text': 'MC'
            },
            'system': 'http://ns.electronichealth.net.au/id/medicare-number',
            'value': medicareNo
        })
    if dvaNo is not None:
        patient_dict['identifier'].append({
            'type': {
                'coding': [
                    {'system': 'http://terminology.hl7.org/CodeSystem/v2-0203',
                    'code': dvaCode[dvaType],
                    'display': 'DVA ' + dvaName[dvaType] + 'Number'}
                ],
                'text': 'DVA Number (' + dvaColor[dvaType] + ')'
            },
            'system': 'http://ns.electronichealth.net.au/id/dva',
            'value': dvaNo
        })
    if PEN is not None:
        patient_dict['identifier'].append({
            'type': {
                'coding': [
                    {'system': 'http://terminology.hl7.org/CodeSystem/v2-0203',
                    'code': 'MC',
                    'display': "Patient's Pensioner Concession Number"}
                ],
                'text': 'PEN'
            },
            'system': 'http://ns.electronichealth.net.au/id/centrelink-customer-reference-number',
            'value': PEN
        })
    if SEN is not None:
        patient_dict['identifier'].append({
            'type': {
                'coding': [
                    {'system': 'http://terminology.hl7.org/CodeSystem/v2-0203',
                    'code': 'MC',
                    'display': "Patient's Senior's Healthcare Concession Number"}
                ],
                'text': 'PEN'
            },
            'system': 'http://ns.electronichealth.net.au/id/centrelink-customer-reference-number',
            'value': SEN
        })
    if HC is not None:
        patient_dict['identifier'].append({
            'type': {
                'coding': [
                    {'system': 'http://terminology.hl7.org/CodeSystem/v2-0203',
                    'code': 'MC',
                    'display': "Patient's Healthcare Card Number"}
                ],
                'text': 'HC'
            },
            'system': 'http://ns.electronichealth.net.au/id/centrelink-customer-reference-number',
            'value': HC
        })
    for GP in patientGPs:
        patient_dict['generalPractitioner'].append({
            'reference': 'PractitionerRole/' + GP 
        })

    patient = Patient.parse_obj(patient_dict)
    try:
        patient_json = json.dumps(patient_dict)
    except Exception as e:
        logging.critical('Bad JSON:%s', e.args)
        sys.stdout.flush()
        logging.shutdown()
        sys.exit(EX_DATAERR)
    # print(json.dumps(patient_dict, indent=4))

    # Return the resources
    if len(patient_json) > 32767:
        logging.critical('Maximum string length for an Excel cell exceeded')
        print(patient_dict)
        sys.stdout.flush()
        logging.shutdown()
        sys.exit(EX_DATAERR)

    return patient_json


if __name__ == '__main__':
    '''
The main code
    '''

    # Save the program name
    progName = sys.argv[0]
    progName = progName[0:-3]        # Strip off the .py ending

    parser = argparse.ArgumentParser()
    parser.add_argument('-D', '--dataDir', dest='dataDir', default='data',
                        help='The name of the directory containing source names and address data(default="data")')
    parser.add_argument('-A', '--addressFile', dest='addressFile', default='GNAF_CORE.psv',
                        help='The file of GNAF_CORE addresses (or subset) (default="GNAF_CORE.psv})')
    parser.add_argument('-O', '--outputDir', dest='outputDir', default='output',
                        help='The name of the output directory [mkHealthPopulation.cfg will be read from this directory] (default="output")')
    parser.add_argument('-o', '--outfile', metavar='outputfile', dest='outputfile', default='healthPopulation.xlsx',
                        help='The name of health population Excel file to be created (default=healthPopulation.xlsx)')
    parser.add_argument('-r', '--makeRandom', dest='makeRandom', action='store_true',
                        help='Make up Australian addresses by random selection of street, suburb, state, postcode')
    parser.add_argument('-P', '--Patients', dest='Patients',
                        action='store_true', help='Output Patients for each doctor')
    parser.add_argument('-i', '--HPI', dest='HPI', action='store_true',
                        help='Add Australian HPI-I, providerNo, prescriberNo and HPI-O numbers')
    parser.add_argument('-x', '--extendNames', dest='extendNames',
                        action='store_true', help='Extend names with sequential letters')
    parser.add_argument('-a', '--addUR', dest='addUR',
                        action='store_true', help='Add a template for the UR number in PID and LIS2 segments')
    parser.add_argument('-v', '--verbose', dest='loggingLevel', type=int, choices=range(0, 5),
                        help='The level of logging\n\t0=CRITICAL,1=ERROR,2=WARNING,3=INFO,4=DEBUG')
    parser.add_argument('-L', '--logDir', dest='logDir', default='logs',
                        help='The name of a directory for the logging file(default="logs")')
    parser.add_argument('-l', '--logfile', metavar='logfile',
                        dest='logfile', help='The name of a logging file')
    args = parser.parse_args()

    # Set up logging
    logging_levels = {0: logging.CRITICAL, 1: logging.ERROR, 2: logging.WARNING, 3: logging.INFO, 4: logging.DEBUG}
    logfmt = progName + ' [%(asctime)s]: %(message)s'
    if args.loggingLevel:   # Change the logging level from "WARN" as the -v vebose option was specified
        loggingLevel = args.loggingLevel
        if args.logfile:        # and send it to a file if the -o logfile option is specified
            logging.basicConfig(format=logfmt, datefmt='%d/%m/%y %H:%M:%S %p', level=logging_levels[loggingLevel],
                                filemode='w', filename=os.path.join(args.logDir, args.logfile))
        else:
            logging.basicConfig(format=logfmt, datefmt='%d/%m/%y %H:%M:%S %p', level=logging_levels[loggingLevel])
    else:                   # send the default (WARN) logging to a file if the -o logfile option is specified
        if args.logfile:
            logging.basicConfig(format=logfmt, datefmt='%d/%m/%y %H:%M:%S %p', filemode='w',
                                filename=os.path.join(args.logDir, args.logfile))
        else:
            logging.basicConfig(format=logfmt, datefmt='%d/%m/%y %H:%M:%S %p')

    # Parse the remaining command line options
    dataDir = args.dataDir
    addressFile = args.addressFile
    outputDir = args.outputDir
    outputfile = args.outputfile
    makeRandom = args.makeRandom
    Patients = args.Patients
    HPI = args.HPI
    extendNames = args.extendNames
    addUR = args.addUR

    # Then read in the configuration from mkHealthPopulation.cfg
    config = configparser.ConfigParser(allow_no_value=True)
    config.optionxform = str
    try:
        config.read(os.path.join(outputDir, 'mkHealthPopulation.cfg'))
        minNetworks = config.getint('Networks', 'minNetworks')
        maxNetworks = config.getint('Networks', 'maxNetworks') + 1
        for row in csv.reader([config.get('Networks', 'networkNames')], csv.excel):
            networkNames = row
            break
        hospitals = {}
        hospitals['associated'] = {}
        hospitals['associated']['minHospitals'] = config.getint('AssociatedHospitals', 'minHospitals')
        hospitals['associated']['maxHospitals'] = config.getint('AssociatedHospitals', 'maxHospitals') + 1
        for row in csv.reader([config.get('AssociatedHospitals', 'hospitalNames')], csv.excel):
            hospitals['associated']['hospitalNames'] = row
            break
        for row in csv.reader([config.get('AssociatedHospitals', 'hospitalPostcodes')], csv.excel):
            hospitals['associated']['hospitalPostcodes'] = row
            break
        for row in csv.reader([config.get('AssociatedHospitals', 'departments')], csv.excel):
            hospitals['associated']['departments'] = row
            break
        hospitals['associated']['minSpecialists'] = config.getint('AssociatedHospitals', 'minSpecialists')
        hospitals['associated']['maxSpecialists'] = config.getint('AssociatedHospitals', 'maxSpecialists') + 1
        hospitals['associated']['minNurses'] = config.getint('AssociatedHospitals', 'minNurses')
        hospitals['associated']['maxNurses'] = config.getint('AssociatedHospitals', 'maxNurses') + 1
        hospitals['private'] = {}
        hospitals['private']['minHospitals'] = config.getint('PrivateHospitals', 'minHospitals')
        hospitals['private']['maxHospitals'] = config.getint('PrivateHospitals', 'maxHospitals') + 1
        hospitals['private']['departments'] = config.get('PrivateHospitals', 'departments')
        for row in csv.reader([config.get('PrivateHospitals', 'hospitalNames')], csv.excel):
            hospitals['private']['hospitalNames'] = row
            break
        for row in csv.reader([config.get('PrivateHospitals', 'hospitalPostcodes')], csv.excel):
            hospitals['private']['hospitalPostcodes'] = row
            break
        for row in csv.reader([config.get('PrivateHospitals', 'departments')], csv.excel):
            hospitals['private']['departments'] = row
            break
        hospitals['private']['minSpecialists'] = config.getint('PrivateHospitals', 'minSpecialists')
        hospitals['private']['maxSpecialists'] = config.getint('PrivateHospitals', 'maxSpecialists') + 1
        hospitals['private']['minNurses'] = config.getint('PrivateHospitals', 'minNurses')
        hospitals['private']['maxNurses'] = config.getint('PrivateHospitals', 'maxNurses') + 1
        minClinics = config.getint('Clinics', 'minClinics')
        maxClinics = config.getint('Clinics', 'maxClinics') + 1
        minConsultants = config.getint('Clinics', 'minConsultants')
        maxConsultants = config.getint('Clinics', 'maxConsultants') + 1
        minDr = config.getint('Clinics', 'minDr')
        maxDr = config.getint('Clinics', 'maxDr') + 1
        minPatients = config.getint('Clinics', 'minPatients')
        maxPatients = config.getint('Clinics', 'maxPatients') + 1
        minAge = config.getint('Patients', 'minAge')
        maxAge = config.getint('Patients', 'maxAge') + 1

        departmentSpecialties = dict(config['DepartmentSpecialties'])
        departmentSpecialist = dict(config['DepartmentSpecialist'])
        departmentRoles = dict(config['DepartmentRoles'])
        GPspecialties = dict(config['GPspecialty'])
        SpecialistRoles = dict(config['SpecialistRoles'])
        for specialty in SpecialistRoles:
            for row in csv.reader([SpecialistRoles[specialty]], csv.excel):
                SpecialistRoles[specialty] = row
                break
        Specialties = dict(config['Specialties'])
        Roles = dict(config['Roles'])
        HealthcareRoles = dict(config['HealthcareRoles'])

        addressFields = config.get('Fields', 'addressFields')
        networkFields = config.get('Fields', 'networkFields')
        hospitalFields = config.get('Fields', 'hospitalFields')
        departmentFields = config.get('Fields', 'departmentFields')
        clinicFields = config.get('Fields', 'clinicFields')
        drFields = config.get('Fields', 'drFields')
        patientFields1 = config.get('Fields', 'patientFields1')
        patientFields2 = config.get('Fields', 'patientFields2')

        for row in csv.reader([addressFields], csv.excel):
            addressFields = row
            break
        for row in csv.reader([networkFields], csv.excel):
            networkFields = row
            break
        for row in csv.reader([hospitalFields], csv.excel):
            hospitalFields = row
            break
        for row in csv.reader([departmentFields], csv.excel):
            departmentFields = row
            break
        for row in csv.reader([clinicFields], csv.excel):
            clinicFields = row
            break
        for row in csv.reader([drFields], csv.excel):
            drFields = row
            break
        for row in csv.reader([patientFields1], csv.excel):
            patientFields1 = row
            break
        for row in csv.reader([patientFields2], csv.excel):
            patientFields2 = row
            break
    except (configparser.MissingSectionHeaderError, configparser.NoSectionError,
            configparser.NoOptionError, configparser.ParsingError) as detail:
        logging.fatal('%s', detail)
        logging.shutdown()
        sys.exit(EX_CONFIG)

    noOfNetworks = random.randrange(minNetworks, maxNetworks)
    noOfRecords = noOfNetworks
    maxHospitals = max(hospitals['associated']['maxHospitals'], hospitals['private']['maxHospitals'])
    noOfRecords += maxHospitals * 2
    maxDepartments = max(len(hospitals['associated']['departments']), len(hospitals['private']['departments']))
    noOfRecords += maxHospitals * maxDepartments * 2
    maxSpecialists = max(hospitals['associated']['maxSpecialists'], hospitals['private']['maxSpecialists'])
    maxNurses = max(hospitals['associated']['maxNurses'], hospitals['private']['maxNurses'])
    noOfRecords += noOfNetworks * maxHospitals * maxDepartments * maxSpecialists * 2
    noOfRecords += noOfNetworks * maxHospitals * maxDepartments * maxNurses * 2
    noOfRecords += maxClinics
    noOfRecords += maxClinics * maxConsultants * 2
    noOfRecords += maxClinics * maxDr

    if Patients:
        noOfRecords += maxClinics * maxDr * maxPatients

    UsedIDs = {}
    # Make random patients with long street names
    logging.info('Creating %d demographic records', noOfRecords)
    mkRandPatients(dataDir, addressFile, noOfRecords, extendNames, False, makeRandom, minAge, maxAge, False, UsedIDs, addUR)        # Create enough random patient


    # Create the Networks, hospitals, clinics, specialists, doctors (and patients if required)
    usedProviderNo = {}
    usedAhpraNo = set()
    wb = Workbook()
    healthNetworks = wb.active
    healthNetworks.title = 'Health Networks'
    healthNetworks.append(['network_HPI-O', 'networkName', 'authority', 'streetNo', 'streetName', 'shortStreetType',
                           'suburb', 'state', 'postcode', 'longitude', 'latitude', 'meshblock', 'sa1', 'country', 'businessPhone'])
    publicHospitals = wb.create_sheet('Public Hospitals')
    publicHospitals.append(['network_HPI-O', 'hospital_HPI-O', 'hospitalName', 'streetNo', 'streetName', 'shortStreetType',
                            'suburb', 'state', 'postcode', 'longitude', 'latitude', 'meshblock', 'sa1', 'country', 'businessPhone'])
    publicHospitalDepartments = wb.create_sheet('Public Hospital Departments')
    publicHospitalDepartments.append(
        ['hospital_HPI-O', 'department_HPI-O', 'departmentName', 'application', 'departmentSpecialty', 'specialtyDescription', 'businessPhone'])
    publicHospitalStaff = wb.create_sheet('Public Hospital Staff')
    publicHospitalStaff.append(['department_HPI-O', 'staffSpecialty', 'specialtyDescription', 'role', 'roleDescription', 'HPI-I',
                                'providerNo', 'prescriberNo', 'ahpraNo', 'title', 'familyName', 'givenName', 'birthdate', 'sex', 'workMobile', 'businessPhone', 'workEmail'])

    privateHospitals = wb.create_sheet('Private Hospitals')
    privateHospitals.append(['hospital_HPI-O', 'hospitalName', 'authority', 'streetNo', 'streetName', 'shortStreetType',
                             'suburb', 'state', 'postcode', 'longitude', 'latitude', 'meshblock', 'sa1', 'country', 'businessPhone'])
    privateHospitalDepartments = wb.create_sheet('Private Hospital Departments')
    privateHospitalDepartments.append(['hospital_HPI-O', 'department_HPI-O', 'departmentName', 'application', 'departmentSpecialty', 'specialtyDescription', 'businessPhone'])
    privateHospitalStaff = wb.create_sheet('Private Hospital Staff')
    privateHospitalStaff.append(['department_HPI-O', 'staffSpecialty', 'specialtyDescription', 'role', 'roleDescription', 'HPI-I',
                                 'providerNo', 'prescriberNo', 'ahpraNo', 'title', 'familyName', 'givenName', 'birthdate', 'sex', 'workMobile', 'businessPhone', 'workEmail'])

    clinics = wb.create_sheet('GP Clinics')
    clinics.append(['clinic_HPI-O', 'clinicName', 'authority', 'application', 'clinicSpecialty', 'specialtyDescription', 'streetNo', 'streetName', 'shortStreetType',
                    'suburb', 'state', 'postcode', 'longitude', 'latitude', 'meshblock', 'sa1', 'country', 'businessPhone'])
    clinicStaff = wb.create_sheet('GP Clinic Staff')
    clinicStaff.append(['clinic_HPI-O', 'staffSpecialty', 'specialtyDescription', 'role', 'roleDescription', 'HPI-I', 'providerNo',
                        'prescriberNo', 'ahpraNo', 'title', 'familyName', 'givenName', 'birthdate', 'sex', 'workMobile', 'businessPhone', 'workEmail'])

    specialistServices = wb.create_sheet('Specialist Services')
    specialistServices.append(['specialistService_HPI-O', 'specialistServiceName', 'authority', 'application', 'serviceSpecialty', 'specialtyDescription', 'streetNo', 'streetName',
                               'shortStreetType', 'suburb', 'state', 'postcode', 'longitude', 'latitude', 'meshblock', 'sa1', 'country', 'businessPhone'])
    specialists = wb.create_sheet('Specialists')
    specialists.append(['specialistService_HPI-O', 'specialistSpecialty', 'specialtyDescription', 'role', 'roleDescription',
                        'HPI-I', 'providerNo', 'prescriberNo', 'ahpraNo', 'title', 'familyName', 'givenName', 'birthdate', 'sex', 'workMobile', 'businessPhone', 'workEmail'])

    if Patients:
        clinicPatients = wb.create_sheet('Patients')
        clinicPatients.append(['clinic_HPI-O', 'GP_HPI-I', 'IHI', 'CentreLink', 'Pension', 'SeniorsHC', 'HealthCare', 'title', 'familyName', 'givenName', 'birthdate', 'sex', 'streetNo', 'streetName', 'shortStreetType', 'suburb', 'state', 'postcode',
                               'longitude', 'latitude', 'meshblock', 'sa1', 'country', 'mobile', 'homePhone', 'businessPhone', 'email', 'medicareNo', 'dvaNo', 'dvaType', 'height', 'weight', 'waist', 'hips', 'married', 'race'])
        HL7_PID = wb.create_sheet('HL7_PID')
        HL7_PID.append(['IHI', 'PID'])
        LIS2_P = wb.create_sheet('LIS2_P')
        LIS2_P.append(['IHI', 'P'])

    HL7_PRD = wb.create_sheet('HL7_PRD')
    HL7_PRD.append(['HPI-I', 'PRD'])

    FHIR_Organization = wb.create_sheet('FHIR_Organization')
    FHIR_Organization.append(['HPI-O', 'Organization'])
    FHIR_HealthcareService = wb.create_sheet('FHIR_HealthcareService')
    FHIR_HealthcareService.append(['HPI-O', 'HealthcareService'])
    FHIR_Location = wb.create_sheet('FHIR_Location')
    FHIR_Location.append(['HPI-O', 'Location'])
    FHIR_Practitioner = wb.create_sheet('FHIR_Practitioner')
    FHIR_Practitioner.append(['HPI-I', 'Practitioner'])
    FHIR_PractitionerRole = wb.create_sheet('FHIR_PractitionerRole')
    FHIR_PractitionerRole.append(['HPI-I', 'PractitionerRole'])
    FHIR_CareTeam = wb.create_sheet('FHIR_CareTeam')
    FHIR_CareTeam.append(['IHI', 'HPI-O', 'CareTeam'])
    if Patients:
        FHIR_Patient = wb.create_sheet('FHIR_Patient')
        FHIR_Patient.append(['IHI', 'Patient'])

    record = 0
    publicHospitalNo = 0
    privateHospitalNo = 0
    usedSpecServiceNames = set()
    usedClinicNames = set()
    usedPubDrHPIO = set()
    usedPrivDrHPIO = set()
    usedNrsHPIO = set()
    usedIHI = set()
    deptHPIOs = set()
    for thisNetwork in range(noOfNetworks):
        # healthNetworks fields:network_HPI-O,networkName,authority,streetNo,streetName,shortStreetType,suburb,state,postcode,longitude,latitude,meshblock,sa1,country,businessPhone
        outputRow = []
        # Assign HPI-O as organization ID
        IHI = patients[patientKeys[record]]['IHI']
        networkHPIO = IHI[:5] + '2' + IHI[6:-1]  # network HPI-O
        networkHPIO = f'{networkHPIO}{mkLuhn(networkHPIO):d}'
        outputRow.append(networkHPIO)
        if thisNetwork == len(networkNames):
            logging.fatal('Insufficient networkNames on configuration')
            logging.shutdown()
            sys.exit(EX_CONFIG)

        # Network - select an SA4
        networkSA2 = random.choice(list(SA1s))[0:9]
        networkSA4 = networkSA2[:3]
        name = networkNames[thisNetwork]
        outputRow.append(name)
        initials = name.split(' ')
        authority = ''
        for initial in initials:
            authority += initial[0].upper()
        outputRow.append(authority)
        # Start with the address (less postcode)
        for field in addressFields:
            if field == 'postcode':
                # Make the postcode one that is within this SA3
                sa3 = networkSA2[:5]
                outputRow.append(random.choice(list(SA3postcodes[sa3])))
            else:
                outputRow.append(patients[patientKeys[record]][field])
        # And the network specific fields
        for field in networkFields:
            outputRow.append(patients[patientKeys[record]][field])
        healthNetworks.append(outputRow)
        FHIR_Organization.append([networkHPIO, createOrganization(record, networkHPIO, None, name, None, 'network')])

        # Now create some hospitals - some public and some private
        for hospital in ['associated', 'private']:
            noOfHospitals = random.randrange(hospitals[hospital]['minHospitals'], hospitals[hospital]['maxHospitals'])
            hospitals[hospital]['hospitalSA3'] = []
            # public hospital fields:network_HPI-O,hospital_HPI-O,hospitalName,streetNo,streetName,shortStreetType,suburb,state,postcode,longitude,latitude,meshblock,sa1,country,businessPhone
            # private hospital fields:hospital_HPI-O,hospitalName,authority,streetNo,streetName,shortStreetType,suburb,state,postcode,longitude,latitude,meshblock,sa1,country,businessPhone
            for thisHospital in range(noOfHospitals):
                outputRow = []
                if hospital == 'associated':
                    outputRow.append(networkHPIO)
                # Assign HPI-O organization ID
                record += 1
                IHI = patients[patientKeys[record]]['IHI']
                hospital_HPIO = IHI[:5] + '2' + IHI[6:-1]
                hospital_HPIO = f'{hospital_HPIO}{mkLuhn(hospital_HPIO):d}'
                outputRow.append(hospital_HPIO)
                # Assign a name and select a different SA2 in this SA4
                if hospital == 'associated':
                    if publicHospitalNo == len(hospitals[hospital]['hospitalNames']):
                        logging.fatal('Insufficient associated hospitalNames in configuration')
                        logging.shutdown()
                        sys.exit(EX_CONFIG)
                    if publicHospitalNo == len(hospitals[hospital]['hospitalPostcodes']):
                        logging.fatal('Insufficient associated hospitalPostcodes in configuration')
                        logging.shutdown()
                        sys.exit(EX_CONFIG)
                    hospitalName = hospitals[hospital]['hospitalNames'][publicHospitalNo]
                    hospitalPostcode = hospitals[hospital]['hospitalPostcodes'][publicHospitalNo]
                    publicHospitalNo += 1
                else:
                    if privateHospitalNo == len(hospitals[hospital]['hospitalNames']):
                        logging.fatal('Insufficient private hospitalNames in configuration')
                        logging.shutdown()
                        sys.exit(EX_CONFIG)
                    if privateHospitalNo == len(hospitals[hospital]['hospitalPostcodes']):
                        logging.fatal('Insufficient private hospitalPostcodes in configuration')
                        logging.shutdown()
                        sys.exit(EX_CONFIG)
                    hospitalName = hospitals[hospital]['hospitalNames'][privateHospitalNo]
                    hospitalPostcode = hospitals[hospital]['hospitalPostcodes'][privateHospitalNo]
                    privateHospitalNo += 1
                outputRow.append(hospitalName)
                if hospital == 'private':
                    initials = hospitalName.split(' ')
                    authority = ''
                    for initial in initials:
                        authority += initial[0].upper()
                    outputRow.append(authority)
                # Create an address - check that the postcode is in our address file
                hospitalSA3 = None
                for thisSA3 in SA3postcodes:
                    if hospitalPostcode in SA3postcodes[thisSA3]:
                        hospitalSA3 = thisSA3
                        break
                if hospitalSA3 is None:
                    logging.fatal('Hospital postcode (%s) not in address file', hsopitalPostcode)
                    logging.shutdown()
                    sys.exit(EX_CONFIG)
                hospitals[hospital]['hospitalSA3'].append(hospitalSA3)
                thisAddr = mkRandAddress(hospitalSA3, True, makeRandom)
                for field in addressFields:
                    outputRow.append(thisAddr[field])
                for field in hospitalFields:
                    outputRow.append(patients[patientKeys[record]][field])
                if hospital == 'associated':
                    publicHospitals.append(outputRow)
                    FHIR_Organization.append([hospital_HPIO, createOrganization(record, hospital_HPIO, networkHPIO, hospitalName, thisAddr, 'public hospital')])
                    FHIR_Location.append([hospital_HPIO, createLocation(record, hospital_HPIO, hospitalName, thisAddr, networkHPIO)])
                else:
                    privateHospitals.append(outputRow)
                    FHIR_Organization.append([hospital_HPIO, createOrganization(record, hospital_HPIO, None, hospitalName, thisAddr, 'private hospital')])
                    FHIR_Location.append([hospital_HPIO, createLocation(record, hospital_HPIO, hospitalName, thisAddr, None)])

                # Now create some departments and staff for this hospital
                # department fields:hospital_HPI-O,department_HPI-O,departmentName,departmentSpecialty,specialtyDescription
                noOfDepartments = random.randrange(2, len(hospitals[hospital]['departments']))
                for thisDepartment in range(noOfDepartments):
                    department = hospitals[hospital]['departments'][thisDepartment]
                    outputRow = []
                    outputRow.append(hospital_HPIO)
                    record += 1
                    IHI = patients[patientKeys[record]]['IHI']
                    department_HPIO = IHI[:5] + '2' + IHI[6:-1]
                    department_HPIO = f'{department_HPIO}{mkLuhn(department_HPIO):d}'
                    deptHPIOs.add(department_HPIO)
                    outputRow.append(department_HPIO)
                    outputRow.append(department)
                    if department in ['Emergency', 'Medical', 'Midwifery', 'Surgical', 'Paediatric', 'IntensiveCare']:
                        outputRow.append('PAS')
                    else:
                        outputRow.append(department[:4].upper())
                    outputRow.append(departmentSpecialties[department])
                    outputRow.append(Specialties[departmentSpecialties[department]])
                    hospitalPhone = patients[patientKeys[record]]['businessPhone']
                    for field in departmentFields:
                        outputRow.append(patients[patientKeys[record]][field])
                    if hospital == 'associated':
                        publicHospitalDepartments.append(outputRow)
                    else:
                        privateHospitalDepartments.append(outputRow)
                    FHIR_Organization.append([department_HPIO, createOrganization(record, department_HPIO, hospital_HPIO, department, thisAddr, 'department')])
                    specialty = departmentSpecialties[department]
                    FHIR_HealthcareService.append([department_HPIO, createHealthcareService(record, department_HPIO, department + '@' + hospitalName, 'department', specialty)])

                    # Now create some staff
                    outputRow = []
                    # staff fields:department_HPI-O,staffSpecialty,specialtyDescription,role,roleDescription,HPI-I,providerNo,prescriberNo,ahpraNo,title,familyName,givenName,birthdate,sex,workMobile,businessPhone,workEmail
                    noOfSpecialists = random.randrange(hospitals[hospital]['minSpecialists'], hospitals[hospital]['maxSpecialists'])
                    deptSpecialists = set()
                    for thisSpecialsist in range(noOfSpecialists):
                        outputRow = []
                        outputRow.append(department_HPIO)
                        outputRow.append(departmentSpecialties[department])
                        outputRow.append(
                            Specialties[departmentSpecialties[department]])
                        role = departmentSpecialist[department]
                        outputRow.append(role)
                        outputRow.append(Roles[role])
                        if hospital == 'associated':
                            if (len(usedPubDrHPIO) > 10) and (len(usedPubDrHPIO) > len(deptSpecialists)) and (random.random() < 0.2):
                                thisRecord = random.choice(list(usedPubDrHPIO))
                                while thisRecord in deptSpecialists:
                                    thisRecord = random.choice(list(usedPubDrHPIO))
                                    logging.debug('Searching for a used public hospital doctor')
                            else:
                                record += 1
                                thisRecord = record
                                usedPubDrHPIO.add(record)
                        else:
                            if (len(usedPrivDrHPIO) > 10) and (len(usedPrivDrHPIO) > len(deptSpecialists)) and (random.random() < 0.2):
                                thisRecord = random.choice(list(usedPrivDrHPIO))
                                while thisRecord in deptSpecialists:
                                    thisRecord = random.choice(list(usedPrivDrHPIO))
                                    logging.debug('Searching for a used private hospital doctor')
                            else:
                                record += 1
                                thisRecord = record
                                usedPrivDrHPIO.add(record)
                        deptSpecialists.add(thisRecord)
                        IHI = patients[patientKeys[thisRecord]]['IHI']
                        specialist_HPII = IHI[:5] + '1' + IHI[6:-1]
                        specialist_HPII = f'{specialist_HPII}{mkLuhn(specialist_HPII):d}'
                        outputRow.append(specialist_HPII)
                        while True:        # Loop if the providerNo is not distinct
                            providerNo = f'{random.randint(100000, 999999):06d}'
                            if providerNo not in usedProviderNo:
                                break
                            logging.debug('Searching for a free provider number for a hospital doctor')
                        usedProviderNo[providerNo] = True
                        outputRow.append(mkProviderNo(providerNo))
                        outputRow.append(mkPrescriberNo(providerNo))
                        while True:        # Loop if the ahpraNo is not distinct
                            ahpraNo = random.randint(9000000000, 9999999999)
                            if ahpraNo not in usedAhpraNo:
                                break
                            logging.debug('Searching for a free AHPRA number for a hospital doctor')
                        usedAhpraNo.add(ahpraNo)
                        outputRow.append('MED' + str(ahpraNo))
                        outputRow.append('Dr.')
                        patients[patientKeys[thisRecord]]['businessPhone'] = hospitalPhone
                        for field in drFields:
                            outputRow.append(patients[patientKeys[thisRecord]][field])
                        if hospital == 'associated':
                            publicHospitalStaff.append(outputRow)
                        else:
                            privateHospitalStaff.append(outputRow)
                        PRD = 'PRD||' + patients[patientKeys[thisRecord]]['familyName'] + '^' + patients[patientKeys[thisRecord]]['givenName'] + '^^^DR^^L'            # PRD-2 Provider Name
                        PRD += '|' + thisAddr['streetNo'] + ' ' + thisAddr['streetName'] + ' ' + thisAddr['streetType']
                        PRD += '^^' + thisAddr['suburb'] + '^' + thisAddr['state'] + '^' + thisAddr['postcode'] + '^AUS^M'   # PRD-3 Provider Address
                        PRD += '||^PRN^PH^^^^^^' + patients[patientKeys[thisRecord]]['homePhone']
                        PRD += '~^PRN^CP^^^^^^' + patients[patientKeys[thisRecord]]['mobile']
                        PRD += '~^NET^Internet^' + patients[patientKeys[thisRecord]]['email']
                        PRD += '~^WPN^PH^^^^^^' + patients[patientKeys[thisRecord]]['businessPhone']            # PRD-5 home phone, mbile, email, business phone
                        PRD += '||' + mkProviderNo(providerNo) + '^AUSHICPR^UPIN'
                        PRD += '~' + mkPrescriberNo(providerNo) + '^AUSHIC^NPI'                                 # PRD-7 Provider identifiers
                        PRDrow = []
                        PRDrow.append(specialist_HPII)
                        PRDrow.append(PRD)
                        HL7_PRD.append(PRDrow)
                        FHIR_Practitioner.append([specialist_HPII, createPractitioner(thisRecord, department_HPIO, specialist_HPII, providerNo, ahpraNo)])
                        FHIR_PractitionerRole.append([specialist_HPII, createPractitionerRole(thisRecord, department_HPIO, specialist_HPII, role, departmentSpecialties[department])])

                    noOfNurses = random.randrange(hospitals[hospital]['minNurses'], hospitals[hospital]['maxNurses'])
                    deptNurses = set()
                    for thisNurse in range(noOfNurses):
                        outputRow = []
                        outputRow.append(department_HPIO)
                        outputRow.append(departmentSpecialties[department])
                        outputRow.append(Specialties[departmentSpecialties[department]])
                        role = '106292003'     # Nurse
                        outputRow.append(role)
                        outputRow.append(Roles[role])
                        if (len(usedNrsHPIO) > 10) and (len(usedNrsHPIO) > len(deptNurses)) and (random.random() < 0.1):
                            thisRecord = random.choice(list(usedNrsHPIO))
                            while thisRecord in deptNurses:
                                thisRecord = random.choice(list(usedNrsHPIO))
                                logging.debug('Searching for a used hospital nurse')
                        else:
                            record += 1
                            thisRecord = record
                            usedNrsHPIO.add(record)
                        deptNurses.add(thisRecord)
                        IHI = patients[patientKeys[thisRecord]]['IHI']
                        nurse_HPII = IHI[:5] + '1' + IHI[6:-1]
                        nurse_HPII = f'{nurse_HPII}{mkLuhn(nurse_HPII):d}'
                        outputRow.append(nurse_HPII)
                        while True:        # Loop if the providerNo is not distinct
                            providerNo = f'{random.randint(100000, 999999):06d}'
                            if providerNo not in usedProviderNo:
                                break
                            logging.debug('Searching for a free provider number for a hospital nurse')
                        usedProviderNo[providerNo] = True
                        outputRow.append(mkProviderNo(providerNo))
                        outputRow.append(mkPrescriberNo(providerNo))
                        while True:        # Loop if the ahpraNo is not distinct
                            ahpraNo = random.randint(9000000000, 9999999999)
                            if ahpraNo not in usedAhpraNo:
                                break
                            logging.debug('Searching for a free AHPRA number for a hospital nurse')
                        usedAhpraNo.add(ahpraNo)
                        outputRow.append('MED' + str(ahpraNo))
                        outputRow.append('Nrs.')
                        patients[patientKeys[thisRecord]]['businessPhone'] = hospitalPhone
                        for field in drFields:
                            outputRow.append(patients[patientKeys[thisRecord]][field])
                        if hospital == 'associated':
                            publicHospitalStaff.append(outputRow)
                        else:
                            privateHospitalStaff.append(outputRow)
                        PRD = 'PRD||' + patients[patientKeys[thisRecord]]['familyName'] + '^' + patients[patientKeys[thisRecord]]['givenName'] + '^^^DR^^L'            # PRD-2 Provider Name
                        PRD += '|' + thisAddr['streetNo'] + ' ' + thisAddr['streetName'] + ' ' + thisAddr['streetType']
                        PRD += '^^' + thisAddr['suburb'] + '^' + thisAddr['state'] + '^' + thisAddr['postcode'] + '^AUS^M'   # PRD-3 Provider Address
                        PRD += '||^PRN^PH^^^^^^' + patients[patientKeys[thisRecord]]['homePhone']
                        PRD += '~^PRN^CP^^^^^^' + patients[patientKeys[thisRecord]]['mobile']
                        PRD += '~^NET^Internet^' + patients[patientKeys[thisRecord]]['email']
                        PRD += '~^WPN^PH^^^^^^' + patients[patientKeys[thisRecord]]['businessPhone']            # PRD-5 home phone, mbile, email, business phone
                        PRD += '||' + mkProviderNo(providerNo) + '^AUSHICPR^UPIN'
                        PRD += '~' + mkPrescriberNo(providerNo) + '^AUSHIC^NPI'                                 # PRD-7 Provider identifiers
                        PRDrow = []
                        PRDrow.append(nurse_HPII)
                        PRDrow.append(PRD)
                        HL7_PRD.append(PRDrow)
                        FHIR_Practitioner.append([nurse_HPII, createPractitioner(thisRecord, department_HPIO, nurse_HPII, providerNo, ahpraNo)])
                        FHIR_PractitionerRole.append([nurse_HPII, createPractitionerRole(thisRecord, department_HPIO, nurse_HPII, role, departmentSpecialties[department])])

        # Now do the local GP clinics for this health network
        noOfClinics = random.randrange(minClinics, maxClinics)
        # clinics fields:clinic_HPI-O,clinicName,authority,application,clinicSpecialty,specialtyDescription,streetNo,streetName,shortStreetType,suburb,state,postcode,longitude,latitude,meshblock,sa1,country,businessPhone,email])
        for thisClinic in range(noOfClinics):
            outputRow = []
            # HPI-O as the organization ID
            record += 1
            IHI = patients[patientKeys[record]]['IHI']
            clinic_HPIO = IHI[:5] + '2' + IHI[6:-1]
            clinic_HPIO = f'{clinic_HPIO}{mkLuhn(clinic_HPIO):d}'
            outputRow.append(clinic_HPIO)
            uniqueName = False
            hospitalSA3 = random.choice(list(hospitals['associated']['hospitalSA3']))
            while not uniqueName:
                # Create a local address close to this hospital
                clinicAddr = mkRandAddress(hospitalSA3, True, makeRandom)
                # Create a clinic from street or suburb plus Medical/Medical Clinic/Medical Centre
                if random.random() < 0.7:            # street or suburb
                    clinicName = clinicAddr['streetName'] + ' ' + clinicAddr['streetType']
                else:
                    clinicName = clinicAddr['suburb']
                clinicName += ' Medical'
                if random.random() < 0.4:            # Clinic or Centre or nothing
                    clinicName += ' Clinic'
                elif random.random() < 0.4:            # Clinic or Centre or nothing
                    clinicName += ' Centre'
                if clinicName in usedClinicNames:
                    logging.debug('Searching of an unused GP clinic name - %s is not unique', clinicName)
                    continue
                usedClinicNames.add(clinicName)
                uniqueName = True
                outputRow.append(clinicName)
            initials = clinicName.split(' ')
            authority = ''
            for initial in initials:
                authority += initial[0].upper()
            outputRow.append(authority)
            outputRow.append('PAS')
            outputRow.append('408443003')
            outputRow.append('General medical practice')
            clinicPhone = patients[patientKeys[record]]['businessPhone']
            for field in addressFields:
                outputRow.append(clinicAddr[field])
            # And the clinic specific fields
            for field in clinicFields:
                outputRow.append(patients[patientKeys[record]][field])
            clinics.append(outputRow)
            FHIR_Organization.append([clinic_HPIO, createOrganization(record, clinic_HPIO, None, clinicName, clinicAddr, 'clinic')])
            FHIR_Location.append([clinic_HPIO, createLocation(record, clinic_HPIO, clinicName, clinicAddr, None)])
            FHIR_HealthcareService.append([clinic_HPIO, createHealthcareService(record, clinic_HPIO, clinicName, 'clinic', '408443003')])

            # Now the local consultants
            noOfConsultants = random.randrange(minConsultants, maxConsultants)
            clinicSpecialists = set()
            for thisConsultant in range(noOfConsultants):
                outputRow = []
                # Consultants a have Specialist Service as their organization
                # specialist services fields:specialistService_HPI-O,specialistServiceName,serviceSpecialty,specialtyDescripion,streetNo,streetName,shortStreetType,suburb,state,postcode,longitude,latitude,meshblock,sa1,country,businessMobile,businessPhone,businessEmail
                record += 1
                IHI = patients[patientKeys[record]]['IHI']
                specialist_HPIO = IHI[:5] + '2' + IHI[6:-1]
                specialist_HPIO = f'{specialist_HPIO}{mkLuhn(specialist_HPIO):d}'
                outputRow.append(specialist_HPIO)
                # Create a local address close to this hospital
                thisAddr = mkRandAddress(hospitalSA3, True, makeRandom)
                # Assign a business name being "suburb specialty services"
                # Pick a specialty
                specialty = random.choice(list(SpecialistRoles))
                # Create the Service Name
                name = thisAddr['suburb'] + ' ' + HealthcareRoles[SpecialistRoles[specialty][1]]
                if name in usedSpecServiceNames:
                    continue
                usedSpecServiceNames.add(name)
                outputRow.append(name)
                initials = name.split(' ')
                authority = ''
                for initial in initials:
                    authority += initial[0].upper()
                outputRow.append(authority)
                outputRow.append('PAS')
                # Add the specialty
                outputRow.append(specialty)
                outputRow.append(Specialties[specialty])
                businessPhone = patients[patientKeys[record]]['businessPhone']
                # Add the address
                for field in addressFields:
                    outputRow.append(thisAddr[field])
                for field in clinicFields:
                    outputRow.append(patients[patientKeys[record]][field])
                specialistServices.append(outputRow)
                FHIR_Organization.append([clinic_HPIO, createOrganization(record, clinic_HPIO, None, name, thisAddr, 'specialist')])
                FHIR_Location.append([specialist_HPIO, createLocation(record, specialist_HPIO, name, thisAddr, None)])
                FHIR_HealthcareService.append([specialist_HPIO, createHealthcareService(record, specialist_HPIO, name, 'specialist', specialty)])

                # Now output the Specialist
                # specialists fields:specialistService_HPI-O,specialistSpecialty,specialtyDescripion,role,roleDescription,HPI-I,providerNo,title,familyName,givenName,birthdate,sex
                for thisSpecialist in range(1, random.randrange(6)):       # Give each specalist clinic between 1 and 5 specialists
                    outputRow = []
                    outputRow.append(specialist_HPIO)
                    outputRow.append(specialty)
                    outputRow.append(Specialties[specialty])
                    role = SpecialistRoles[specialty][0]
                    outputRow.append(role)
                    outputRow.append(Roles[role])
                    if (len(usedPrivDrHPIO) > 10) and (len(usedPrivDrHPIO) > len(clinicSpecialists)) and (random.random() < 0.3):
                        thisRecord = random.choice(list(usedPrivDrHPIO))
                        while thisRecord in clinicSpecialists:
                            thisRecord = random.choice(list(usedPrivDrHPIO))
                            logging.debug('Searching for a used clinicSpecialist')
                    else:
                        record += 1
                        thisRecord = record
                    clinicSpecialists.add(thisRecord)
                    IHI = patients[patientKeys[thisRecord]]['IHI']
                    specialist_HPII = IHI[:5] + '1' + IHI[6:-1]
                    specialist_HPII = f'{specialist_HPII}{mkLuhn(specialist_HPII):d}'
                    outputRow.append(specialist_HPII)
                    while True:        # Loop if the providerNo is not distinct
                        providerNo = f'{random.randint(100000, 999999):06d}'
                        if providerNo not in usedProviderNo:
                            break
                        logging.debug('Searching for a free provider number for a specialist')
                    usedProviderNo[providerNo] = True
                    outputRow.append(mkProviderNo(providerNo))
                    outputRow.append(mkPrescriberNo(providerNo))
                    while True:        # Loop if the ahpraNo is not distinct
                        ahpraNo = random.randint(9000000000, 9999999999)
                        if ahpraNo not in usedAhpraNo:
                            break
                        logging.debug('Searching for a free AHPRA number for a specialist')
                    usedAhpraNo.add(ahpraNo)
                    outputRow.append('MED' + str(ahpraNo))
                    if patients[patientKeys[thisRecord]]['sex'] == 'F':
                        title = 'Ms.'
                    else:
                        title = 'Mr.'
                    outputRow.append(title)
                    patients[patientKeys[thisRecord]]['businessPhone'] = businessPhone
                    # And the Dr fields
                    for field in drFields:
                        outputRow.append(patients[patientKeys[thisRecord]][field])
                    specialists.append(outputRow)
                    PRD = 'PRD||' + patients[patientKeys[thisRecord]]['familyName'] + '^' + patients[patientKeys[thisRecord]]['givenName'] + '^^^DR^^L'            # PRD-2 Provider Name
                    PRD += '|' + thisAddr['streetNo'] + ' ' + thisAddr['streetName'] + ' ' + thisAddr['streetType']
                    PRD += '^^' + thisAddr['suburb'] + '^' + thisAddr['state'] + '^' + thisAddr['postcode'] + '^AUS^M'   # PRD-3 Provider Address
                    PRD += '||^PRN^PH^^^^^^' + patients[patientKeys[thisRecord]]['homePhone']
                    PRD += '~^PRN^CP^^^^^^' + patients[patientKeys[thisRecord]]['mobile']
                    PRD += '~^NET^Internet^' + patients[patientKeys[thisRecord]]['email']
                    PRD += '~^WPN^PH^^^^^^' + patients[patientKeys[thisRecord]]['businessPhone']            # PRD-5 home phone, mbile, email, business phone
                    PRD += '||' + mkProviderNo(providerNo) + '^AUSHICPR^UPIN'
                    PRD += '~' + mkPrescriberNo(providerNo) + '^AUSHIC^NPI'                                 # PRD-7 Provider identifiers
                    PRDrow = []
                    PRDrow.append(specialist_HPII)
                    PRDrow.append(PRD)
                    HL7_PRD.append(PRDrow)
                    FHIR_Practitioner.append([specialist_HPII, createPractitioner(thisRecord, specialist_HPIO, specialist_HPII, providerNo, ahpraNo)])
                    FHIR_PractitionerRole.append([specialist_HPII, createPractitionerRole(thisRecord, specialist_HPIO, specialist_HPII, role, specialty)])

            # Now the clinic doctors and their patients
            noOfDoctors = random.randrange(minDr, maxDr)
            clinicDoctors = set()
            for thisDoctor in range(noOfDoctors):
                # clinic Staff fields:clinic_HPI-O,staffSpecialty,specialtyDescripion,role,roleDescription,HPI-I,providerNo,title,familyName,givenName,birthdate,sex,workMobile,businessPhone,workEmail
                outputRow = []
                outputRow.append(clinic_HPIO)
                # Pick a specialty
                specialty = random.choice(list(GPspecialties))
                outputRow.append(specialty)
                outputRow.append(Specialties[specialty])
                role = GPspecialties[specialty]
                outputRow.append(role)
                outputRow.append(Roles[role])
                # Doctors have their own HPI-I codes
                if (len(usedPubDrHPIO) > 10) and (len(usedPubDrHPIO) > len(clinicDoctors)) and (random.random() < 0.3):
                    thisRecord = random.choice(list(usedPubDrHPIO))
                    while thisRecord in clinicDoctors:
                        thisRecord = random.choice(list(usedPubDrHPIO))
                        logging.debug('Searching for a used public doctor')
                else:
                    record += 1
                    thisRecord = record
                clinicDoctors.add(thisRecord)
                IHI = patients[patientKeys[thisRecord]]['IHI']
                GP_HPII = IHI[:5] + '1' + IHI[6:-1]
                GP_HPII = f'{GP_HPII}{mkLuhn(GP_HPII):d}'
                outputRow.append(GP_HPII)
                while True:        # Loop if the providerNo is not distinct
                    providerNo = f'{random.randint(100000, 999999):06d}'
                    if providerNo not in usedProviderNo:
                        break
                    logging.debug('Searching for a free provider number for a GP')
                usedProviderNo[providerNo] = True
                outputRow.append(mkProviderNo(providerNo))
                outputRow.append(mkPrescriberNo(providerNo))
                while True:        # Loop if the ahpraNo is not distinct
                    ahpraNo = random.randint(9000000000, 9999999999)
                    if ahpraNo not in usedAhpraNo:
                        break
                    logging.debug('Searching for a free AHPRA number for a GP')
                usedAhpraNo.add(ahpraNo)
                outputRow.append('MED' + str(ahpraNo))
                outputRow.append('Dr.')
                patients[patientKeys[thisRecord]]['businessPhone'] = clinicPhone
                # And the Dr fields
                for field in drFields:
                    outputRow.append(patients[patientKeys[thisRecord]][field])
                clinicStaff.append(outputRow)
                PRD = 'PRD||' + patients[patientKeys[thisRecord]]['familyName'] + '^' + patients[patientKeys[thisRecord]]['givenName'] + '^^^DR^^L'            # PRD-2 Provider Name
                PRD += '|' + clinicAddr['streetNo'] + ' ' + clinicAddr['streetName'] + ' ' + clinicAddr['streetType']
                PRD += '^^' + clinicAddr['suburb'] + '^' + clinicAddr['state'] + '^' + clinicAddr['postcode'] + '^AUS^M'   # PRD-3 Provider Address
                PRD += '||^PRN^PH^^^^^^' + patients[patientKeys[thisRecord]]['homePhone']
                PRD += '~^PRN^CP^^^^^^' + patients[patientKeys[thisRecord]]['mobile']
                PRD += '~^NET^Internet^' + patients[patientKeys[thisRecord]]['email']
                PRD += '~^WPN^PH^^^^^^' + patients[patientKeys[thisRecord]]['businessPhone']            # PRD-5 home phone, mbile, email, business phone
                PRD += '||' + mkProviderNo(providerNo) + '^AUSHICPR^UPIN'
                PRD += '~' + mkPrescriberNo(providerNo) + '^AUSHIC^NPI'                                 # PRD-7 Provider identifiers
                PRDrow = []
                PRDrow.append(GP_HPII)
                PRDrow.append(PRD)
                HL7_PRD.append(PRDrow)
                FHIR_Practitioner.append([GP_HPII, createPractitioner(thisRecord, clinic_HPIO, GP_HPII, providerNo, ahpraNo)])
                FHIR_PractitionerRole.append([GP_HPII, createPractitionerRole(thisRecord, clinic_HPIO, GP_HPII, role, specialty)])

                # Save the CareTeams - the hospital one's may be needed on admission


                # An finally the patients for this GP
                if Patients:
                    # patients fields:clinic_HPI-O,GP_HPI-I,IHI,title,familyName,givenName,birthdate,sex,streetNo,streetName,shortStreetType,suburb,state,postcode,longitude,latitude,meshblock,sa1,country,mobile,homePhone,businessPhone,email,medicareNo,dvaNo,dvaType,height,weight,waist,hips,married,race
                    noOfPatients = random.randrange(minPatients, maxPatients)
                    clinicPatientRecords = set()
                    for thisPatient in range(noOfPatients):
                        if (len(usedIHI) > 20) and (len(usedIHI) > len(clinicPatientRecords)) and (random.random() < 0.1):
                            thisRecord = random.choice(list(usedIHI))
                            while thisRecord in clinicPatientRecords:
                                thisRecord = random.choice(list(usedIHI))
                                logging.debug('Searching for a used patient')
                            IHI = patients[patientKeys[thisRecord]]['IHI']
                        else:
                            record += 1
                            thisRecord = record
                            usedIHI.add(record)
                            IHI = patients[patientKeys[record]]['IHI']
                            PIDrow = []
                            PIDrow.append(IHI)
                            thisPID = patients[patientKeys[record]]['PID']
                            PIDrow.append(thisPID)
                            HL7_PID.append(PIDrow)
                            Prow = []
                            Prow.append(IHI)
                            thisLIS2 = patients[patientKeys[record]]['LIS2']
                            Prow.append(thisLIS2)
                            LIS2_P.append(Prow)
                            patientDetails[IHI] = {}
                            patientDetails[IHI]['record'] = thisRecord
                            patientDetails[IHI]['GPs'] = []
                            clinicPatientRecords.add(thisRecord)
                            outputRow = []
                            outputRow.append(clinic_HPIO)
                            outputRow.append(GP_HPII)
                            for field in patientFields1:
                                outputRow.append(patients[patientKeys[thisRecord]][field])
                            for field in addressFields:
                                outputRow.append(patients[patientKeys[thisRecord]][field])
                            for field in patientFields2:
                                outputRow.append(patients[patientKeys[thisRecord]][field])
                            clinicPatients.append(outputRow)
                        patientDetails[IHI]['GPs'].append(GP_HPII + '-' + clinic_HPIO)
                        FHIR_CareTeam.append([IHI, clinic_HPIO, createCareTeam(thisRecord, IHI, clinic_HPIO)])

    # Then the patients details
    if Patients:
        for IHI, IHIdetails in patientDetails.items():
            thisRecord = IHIdetails['record']
            GPs = IHIdetails['GPs']
            FHIR_Patient.append([IHI, createPatient(thisRecord, IHI, GPs)])

    # Then template for department care teams
    for deptHPIO in deptHPIOs:
        FHIR_CareTeam.append(['ReplaceWithIHI', deptHPIO, createCareTeam(None, 'ReplaceWithIHI', deptHPIO)])

    wb.save(os.path.join(outputDir, outputfile))
    logging.shutdown()
    sys.exit(EX_OK)
